import openpyxl

def dataGenerator():

    wk = openpyxl.load_workbook("../excelsheet/Login.xlsx")
    sh = wk["Sheet1"]
    rows = sh.max_row
    cols = sh.max_column
    l1 =[]
    outerlist =[]

    for i in range(2,rows+1):
        l1 =[]
        uname = sh.cell(i,1)
        pwd   = sh.cell(i,2)
        l1.insert(0,uname.value)
        l1.insert(1, pwd.value)
        print(l1)
        outerlist.insert(i-1,l1)

    return outerlist





